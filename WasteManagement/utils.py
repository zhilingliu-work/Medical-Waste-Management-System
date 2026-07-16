# WasteManagement/utils.py - Department Management Utilities
import csv
import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from django.http import HttpResponse
from .models import Department, WasteType, WasteRecord


class DepartmentDataExporter:
    """Utility class for exporting department data"""

    @staticmethod
    def export_to_csv(year: int, month: Optional[int] = None) -> HttpResponse:
        """Export department data to Excel format"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            import pandas as pd
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return DepartmentDataExporter._export_to_csv_fallback(year, month)

        # Build filename
        if month:
            filename = f"部門廢棄物資料_{year}年{month:02d}月.xlsx"
            date_filter = f"{year}-{month:02d}"
        else:
            filename = f"部門廢棄物資料_{year}年.xlsx"
            date_filter = f"{year}"

        # Get data
        if month:
            records = WasteRecord.objects.filter(date=date_filter).select_related('department')
            dates = [date_filter]
        else:
            records = WasteRecord.objects.filter(date__startswith=date_filter).select_related('department')
            dates = sorted(list(set(record.date for record in records)))

        departments = Department.objects.filter(is_active=True).order_by('display_order', 'name')

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "部門廢棄物資料"

        # Write header
        headers = ['日期'] + [dept.name for dept in departments]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        if not dates:
            ws.cell(row=2, column=1, value='無資料')
        else:
            for row_idx, date in enumerate(dates, 2):
                ws.cell(row=row_idx, column=1, value=date)
                date_records = {r.department_id: r.amount for r in records.filter(date=date)}

                for col_idx, dept in enumerate(departments, 2):
                    amount = date_records.get(dept.id)
                    if amount is not None:
                        ws.cell(row=row_idx, column=col_idx, value=amount)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value='')

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response

    @staticmethod
    def _export_to_csv_fallback(year: int, month: Optional[int] = None) -> HttpResponse:
        """Fallback CSV export if Excel library not available"""
        # Build filename
        if month:
            filename = f"部門廢棄物資料_{year}年{month:02d}月.csv"
            date_filter = f"{year}-{month:02d}"
        else:
            filename = f"部門廢棄物資料_{year}年.csv"
            date_filter = f"{year}"

        # Create CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Add BOM for Excel compatibility
        response.write('\ufeff')

        writer = csv.writer(response)

        # Get data
        if month:
            records = WasteRecord.objects.filter(date=date_filter).select_related('department')
        else:
            records = WasteRecord.objects.filter(date__startswith=date_filter).select_related('department')

        if not records.exists():
            writer.writerow(['日期', '部門', '數量'])
            writer.writerow(['', '無資料', ''])
            return response

        # Group by date
        dates = sorted(list(set(record.date for record in records)))
        departments = Department.objects.filter(is_active=True).order_by('display_order', 'name')

        # Write header
        header = ['日期'] + [dept.name for dept in departments]
        writer.writerow(header)

        # Write data rows
        for date in dates:
            row = [date]
            date_records = {r.department_id: r.amount for r in records.filter(date=date)}

            for dept in departments:
                amount = date_records.get(dept.id, '')
                if amount is not None:
                    row.append(amount)
                else:
                    row.append('')

            writer.writerow(row)

        return response

    @staticmethod
    def export_summary_report(year: int) -> Dict:
        """Generate summary report for a year"""
        records = WasteRecord.objects.filter(
            date__startswith=str(year)
        ).select_related('department', 'waste_type')

        if not records.exists():
            return {
                'year': year,
                'total_records': 0,
                'departments_with_data': 0,
                'monthly_totals': {},
                'department_totals': {}
            }

        # Calculate summaries
        monthly_totals = {}
        department_totals = {}

        for record in records:
            if record.amount is not None:
                # Monthly totals
                month = record.date.split('-')[1]
                monthly_totals[month] = monthly_totals.get(month, 0) + record.amount

                # Department totals
                dept_name = record.department.name
                department_totals[dept_name] = department_totals.get(dept_name, 0) + record.amount

        return {
            'year': year,
            'total_records': records.count(),
            'departments_with_data': len(department_totals),
            'monthly_totals': monthly_totals,
            'department_totals': department_totals,
            'grand_total': sum(department_totals.values())
        }


class DepartmentDataValidator:
    """Utility class for validating department data"""

    @staticmethod
    def validate_csv_structure(csv_content: str) -> Tuple[bool, str, List[str]]:
        """Validate CSV file structure and return headers if valid"""
        try:
            # Parse CSV
            csv_reader = csv.reader(io.StringIO(csv_content))
            headers = next(csv_reader, [])

            if not headers:
                return False, "CSV檔案為空", []

            # Check for date column
            if '日期' not in headers:
                return False, "CSV檔案必須包含「日期」欄位", headers

            # Check for valid department columns
            from .models import DepartmentWasteConfiguration
            valid_departments = set(DepartmentWasteConfiguration.get_department_mapping().keys())

            dept_columns = [h for h in headers if h != '日期']
            invalid_departments = [h for h in dept_columns if h not in valid_departments]

            if invalid_departments:
                return False, f"未知部門: {', '.join(invalid_departments[:5])}", headers

            return True, "格式正確", headers

        except Exception as e:
            return False, f"CSV格式錯誤: {str(e)}", []

    @staticmethod
    def validate_date_format(date_str: str) -> bool:
        """Validate date format (YYYY-MM)"""
        if not date_str or len(date_str) != 7 or date_str[4] != '-':
            return False

        try:
            year, month = date_str.split('-')
            year = int(year)
            month = int(month)
            return 1970 <= year <= 9999 and 1 <= month <= 12
        except (ValueError, TypeError):
            return False

    @staticmethod
    def validate_amount(amount_str: str) -> Tuple[bool, Optional[float], str]:
        """Validate amount value"""
        if not amount_str or amount_str.strip() == '':
            return True, None, ""  # Empty is valid (null)

        try:
            amount = float(amount_str.strip())
            if amount < 0:
                return False, None, "數量不能為負數"
            return True, amount, ""
        except ValueError:
            return False, None, "無效的數字格式"


class DepartmentStatistics:
    """Utility class for department statistics calculation"""

    @staticmethod
    def get_monthly_trends(department_id: int, start_year: int, end_year: int) -> Dict:
        """Get monthly trends for a specific department"""
        records = WasteRecord.objects.filter(
            department_id=department_id,
            date__gte=f"{start_year}-01",
            date__lte=f"{end_year}-12"
        ).order_by('date')

        trends = {}
        for record in records:
            if record.amount is not None:
                year, month = record.date.split('-')
                month_key = f"{year}-{month}"
                trends[month_key] = record.amount

        return trends

    @staticmethod
    def get_department_rankings(date: str, limit: int = 10) -> List[Dict]:
        """Get top departments by waste amount for a specific month"""
        records = WasteRecord.objects.filter(
            date=date,
            amount__isnull=False
        ).select_related('department').order_by('-amount')[:limit]

        return [
            {
                'department_name': record.department.name,
                'amount': record.amount,
                'rank': idx + 1
            }
            for idx, record in enumerate(records)
        ]

    @staticmethod
    def calculate_growth_rate(current_amount: float, previous_amount: float) -> Optional[float]:
        """Calculate growth rate between two periods"""
        if previous_amount == 0:
            return None  # Cannot calculate growth from zero

        return ((current_amount - previous_amount) / previous_amount) * 100


def get_date_range_for_year(year: int) -> Tuple[str, str]:
    """Get start and end date strings for a year"""
    return f"{year}-01", f"{year}-12"


def format_amount_display(amount: Optional[float], unit: str = 'metric_ton') -> str:
    """Format amount for display with appropriate precision"""
    if amount is None:
        return '無資料'

    # Format based on unit and value
    if unit == 'new_taiwan_dollar':
        return f"{int(amount):,}"
    elif amount == int(amount):
        return str(int(amount))
    else:
        return f"{amount:.2f}".rstrip('0').rstrip('.')


def parse_month_input(month_input: str) -> Optional[Tuple[int, int]]:
    """Parse month input (YYYY-MM) into year and month integers"""
    if not month_input or len(month_input) != 7:
        return None

    try:
        year, month = month_input.split('-')
        return int(year), int(month)
    except (ValueError, TypeError):
        return None